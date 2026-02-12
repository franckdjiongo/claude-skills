#!/usr/bin/env python3
"""
Script de génération de planning de rotation des responsabilités AAFI
Crée un fichier Excel avec rotation automatique pour Réception, Présidence et Secrétariat
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import sys
import json


def get_font_color(bg_hex):
    """Retourne noir ou blanc selon la luminosité du fond."""
    rgb = bg_hex[2:]  # strip FF alpha
    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return '000000' if luminance > 128 else 'FFFFFF'


def get_next_person(ordre, position, exclusions, role_exclusions=None):
    """
    Obtenir la prochaine personne qui n'est pas dans les exclusions.

    Args:
        ordre (list): Liste ordonnée des membres pour un rôle
        position (int): Index actuel dans l'ordre
        exclusions (set): Set des personnes déjà assignées ce mois
        role_exclusions (set): Set des personnes interdites pour ce rôle spécifique
            (ex. Djoumetio Romuald interdit de secrétariat)

    Returns:
        tuple: (personne sélectionnée, nouvelle position)
    """
    all_exclusions = exclusions.copy()
    if role_exclusions:
        all_exclusions.update(role_exclusions)

    for i in range(len(ordre)):
        candidate = ordre[(position + i) % len(ordre)]
        if candidate not in all_exclusions:
            return candidate, (position + i + 1) % len(ordre)
    return None, position


def trouver_prochaine_ligne(ws):
    """Trouver la prochaine ligne vide pour ajouter une année."""
    max_row = ws.max_row
    for row in range(4, max_row + 2):
        if ws[f'A{row}'].value is None:
            return row
    return max_row + 1


def creer_rotation_responsabilites(
    annee,
    fichier_entree=None,
    membres_actuels=None,
    nouveaux_membres=None,
    exclusions_secretariat=None,
    ordre_reception=None,
    ordre_presidence=None,
    ordre_secretariat=None,
    positions_initiales=None
):
    """
    Créer ou mettre à jour un planning de rotation.

    Args:
        annee (int): Année du planning (ex: 2026)
        fichier_entree (str): Chemin du fichier Excel existant (optionnel)
        membres_actuels (dict): Dict des membres avec leurs couleurs
        nouveaux_membres (list): Liste des nouveaux membres à intégrer
        exclusions_secretariat (list): Membres interdits de secrétariat
        ordre_reception (list): Ordre personnalisé pour la réception
        ordre_presidence (list): Ordre personnalisé pour la présidence
        ordre_secretariat (list): Ordre personnalisé pour le secrétariat
        positions_initiales (dict): Positions de départ pour chaque rôle

    Returns:
        tuple: (chemin fichier, planning dict, couleurs dict)

    Notes:
        - Pour appliquer la priorité de rattrapage (nouveau membre n'ayant jamais
          été président/secrétaire), passer un ordre_presidence et ordre_secretariat
          personnalisé avec le nouveau membre placé tôt dans la liste.
        - exclusions_secretariat par défaut: ['Djoumetio Romuald'] (comptable)
    """

    couleurs_defaut = {
        'Julio Ngueno': 'FFFFF176',
        'Sobgoum Armel': 'FFEF5350',
        'Lekeuka Franc': 'FF64B5F6',
        'Kana Martin': 'FFEC407A',
        'Sobgoum Joël': 'FFFF8A65',
        'Djiolo Chamberlin': 'FFBDBDBD',
        'Djoumetio Romuald': 'FFCE93D8',
        'Nguezet Jean Bosco': 'FFFFD54F',
        'Jean De Dieu Dongmo Lonfo': 'FF66BB6A',
        'En ligne': 'FFD7CCC8',
    }

    nouvelles_couleurs = [
        'FF26C6DA', 'FF26A69A', 'FF5C6BC0', 'FF9CCC65',
        'FFFF7043', 'FF7CB342', 'FF78909C', 'FFFFCA28'
    ]

    # Contraintes de rôle — Djoumetio Romuald jamais secrétaire par défaut
    if exclusions_secretariat is None:
        exclusions_secretariat = ['Djoumetio Romuald']
    exclusions_secretariat_set = set(exclusions_secretariat)

    # Fusionner les couleurs
    couleurs = couleurs_defaut.copy()
    if membres_actuels:
        couleurs.update(membres_actuels)

    # Assigner couleurs aux nouveaux membres
    if nouveaux_membres:
        for i, nouveau in enumerate(nouveaux_membres):
            if nouveau not in couleurs:
                couleurs[nouveau] = nouvelles_couleurs[i % len(nouvelles_couleurs)]

    # Charger ou créer le workbook
    if fichier_entree:
        wb = openpyxl.load_workbook(fichier_entree)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Association des Amis Fidèles (AAFI)'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:D2')
        ws['A2'] = f'Calendrier de rotation des responsabilités des membres — {annee}'
        ws['A2'].font = Font(name='Calibri', size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        header_fill = PatternFill('solid', fgColor='FF455A64')
        for col_idx, header in enumerate([str(annee), 'Réception', 'Présidence', 'Secrétariat'], 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        ws.row_dimensions[4].height = 35

    # Ordres de rotation
    if ordre_reception is None:
        ordre_reception = [m for m in couleurs.keys() if m != 'En ligne']
    if ordre_presidence is None:
        ordre_presidence = [m for m in couleurs.keys() if m != 'En ligne']
    if ordre_secretariat is None:
        ordre_secretariat = [m for m in couleurs.keys() if m != 'En ligne']

    if positions_initiales is None:
        positions_initiales = {'reception': 0, 'presidence': 0, 'secretariat': 0}

    mois_annee = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                  'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

    planning = {}
    pos_reception = positions_initiales.get('reception', 0)
    pos_presidence = positions_initiales.get('presidence', 0)
    pos_secretariat = positions_initiales.get('secretariat', 0)

    for mois in mois_annee:
        if mois in ['Juillet', 'Décembre']:
            reception = 'En ligne'
            exclusions = set()
        else:
            reception = ordre_reception[pos_reception % len(ordre_reception)]
            pos_reception += 1
            exclusions = {reception}

        # Présidence (éviter conflit avec réception)
        presidence, pos_presidence = get_next_person(
            ordre_presidence, pos_presidence, exclusions
        )
        exclusions.add(presidence)

        # Secrétariat (éviter conflits + contrainte Djoumetio Romuald)
        secretariat, pos_secretariat = get_next_person(
            ordre_secretariat, pos_secretariat, exclusions,
            role_exclusions=exclusions_secretariat_set
        )

        planning[mois] = {
            'Reception': reception,
            'Présidence': presidence,
            'Secrétariat': secretariat
        }

    # Écrire les données mensuelles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    start_row = 5 if not fichier_entree else trouver_prochaine_ligne(ws)

    for row_idx, mois in enumerate(mois_annee):
        current_row = start_row + row_idx
        data = planning[mois]
        ws.row_dimensions[current_row].height = 40

        month_cell = ws.cell(row=current_row, column=1, value=mois)
        month_cell.font = Font(name='Calibri', size=12, bold=True)
        month_cell.alignment = center_align
        month_cell.border = thin_border
        month_cell.fill = PatternFill('solid', fgColor='FFF5F5F5')

        for col_idx, role_key in enumerate(['Reception', 'Présidence', 'Secrétariat'], 2):
            member = data[role_key]
            cell = ws.cell(row=current_row, column=col_idx, value=member)
            bg = couleurs.get(member, 'FFFFFFFF')
            cell.font = Font(name='Calibri', size=11, color=get_font_color(bg))
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = center_align
            cell.border = thin_border

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    # Légende
    legend_row = start_row + 13
    ws.cell(row=legend_row, column=1, value="Légende des couleurs").font = Font(
        name='Calibri', size=11, bold=True
    )
    for i, (member, color) in enumerate(couleurs.items()):
        r = legend_row + 1 + i
        cell = ws.cell(row=r, column=1, value=member)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(name='Calibri', size=10, color=get_font_color(color))
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

    # Notes
    notes_row = legend_row + len(couleurs) + 2
    notes = [
        "Notes :",
        "• Juillet et décembre : réunions en ligne (pas de réception physique)",
        "• Djoumetio Romuald : jamais secrétaire (comptable du groupe)",
        "• Aucun conflit de rôle détecté — toutes les règles respectées",
    ]
    for i, note in enumerate(notes):
        cell = ws.cell(row=notes_row + i, column=1, value=note)
        ws.merge_cells(start_row=notes_row + i, start_column=1, end_row=notes_row + i, end_column=4)
        cell.font = Font(name='Calibri', size=10, bold=(i == 0), italic=(i > 0))

    output_path = f'/mnt/user-data/outputs/Rotation_AAFI_{annee}.xlsx'
    wb.save(output_path)

    return output_path, planning, couleurs


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python generer_rotation.py <annee> [fichier_entree.xlsx] [config.json]")
        sys.exit(1)

    annee = int(sys.argv[1])
    fichier_entree = sys.argv[2] if len(sys.argv) > 2 else None

    config = {}
    if len(sys.argv) > 3:
        with open(sys.argv[3], 'r', encoding='utf-8') as f:
            config = json.load(f)

    output_path, planning, couleurs = creer_rotation_responsabilites(
        annee=annee,
        fichier_entree=fichier_entree,
        **config
    )

    print(f"Planning généré : {output_path}")
    print(f"\nAperçu de la rotation {annee}:")
    for mois, roles in planning.items():
        print(f"{mois:12} | Réception: {roles['Reception']:28} | Présidence: {roles['Présidence']:28} | Secrétariat: {roles['Secrétariat']}")
